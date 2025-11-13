"""
Экспорт расписания в различные форматы
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from typing import List, Dict, Optional


class ExcelExporter:
    """Класс для экспорта расписания в Excel"""
    
    def __init__(self):
        """Инициализация экспортера"""
        self.days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
        self.times = [
            "08:00-09:30",
            "09:40-11:10",
            "11:20-12:50",
            "13:30-15:00",
            "15:10-16:40",
            "16:50-18:20",
            "18:30-20:00"
        ]
        
        # Стили
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14)
        self.cell_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_schedule(self, schedule, output_path: str, export_type: str = 'group'):
        """
        Экспорт расписания в Excel
        
        Args:
            schedule: Объект расписания из БД
            output_path: Путь для сохранения файла
            export_type: Тип экспорта ('group', 'teacher', 'room', 'consolidated')
        """
        if export_type == 'group':
            self._export_by_groups(schedule, output_path)
        elif export_type == 'teacher':
            self._export_by_teachers(schedule, output_path)
        elif export_type == 'room':
            self._export_by_rooms(schedule, output_path)
        elif export_type == 'consolidated':
            self._export_consolidated(schedule, output_path)
        else:
            self._export_by_groups(schedule, output_path)
    
    def _export_by_groups(self, schedule, output_path: str):
        """Экспорт по группам (отдельный лист для каждой группы)"""
        wb = Workbook()
        
        # Группируем занятия по группам
        groups_lessons = defaultdict(list)
        for lesson in schedule.lessons:
            groups_lessons[lesson.group.name].append(lesson)
        
        # Удаляем лист по умолчанию
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Создаем лист для каждой группы
        for group_name, lessons in sorted(groups_lessons.items()):
            ws = wb.create_sheet(title=group_name[:31])  # Excel ограничение 31 символ
            self._fill_group_sheet(ws, lessons, group_name, schedule)
        
        wb.save(output_path)
    
    def _fill_group_sheet(self, ws, lessons: List, group_name: str, schedule):
        """Заполнение листа для группы"""
        
        # Заголовок
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f'Расписание группы {group_name}'
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Дополнительная информация
        ws['A2'] = f'Семестр: {schedule.semester or "Не указан"}'
        ws['A2'].font = Font(size=10)
        ws['D2'] = f'Учебный год: {schedule.academic_year or "Не указан"}'
        ws['D2'].font = Font(size=10)
        
        # Шапка таблицы
        headers = ['Время'] + self.days[:5]  # Только рабочие дни
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Временные слоты
        for i, time in enumerate(self.times, start=5):
            cell = ws.cell(row=i, column=1)
            cell.value = time
            cell.font = Font(bold=True, size=10)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Создаем матрицу расписания
        schedule_matrix = {}
        for lesson in lessons:
            key = (lesson.day, lesson.time_slot)
            schedule_matrix[key] = lesson
        
        # Заполняем занятия
        for day in range(5):  # Только рабочие дни
            for time_slot in range(7):  # 7 пар
                row = 5 + time_slot
                col = 2 + day
                
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                
                if (day, time_slot) in schedule_matrix:
                    lesson = schedule_matrix[(day, time_slot)]
                    
                    # Форматируем текст занятия
                    cell_value = f"{lesson.subject.name}\n"
                    cell_value += f"👨‍🏫 {lesson.teacher.name}\n"
                    cell_value += f"🏫 ауд. {lesson.room.name}"
                    
                    cell.value = cell_value
                    cell.font = self.cell_font
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 28
        
        # Высота строк
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[4].height = 25
        for row in range(5, 12):
            ws.row_dimensions[row].height = 65
    
    def _export_by_teachers(self, schedule, output_path: str):
        """Экспорт по преподавателям"""
        wb = Workbook()
        
        teachers_lessons = defaultdict(list)
        for lesson in schedule.lessons:
            teachers_lessons[lesson.teacher.name].append(lesson)
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        for teacher_name, lessons in sorted(teachers_lessons.items()):
            ws = wb.create_sheet(title=teacher_name[:31])
            self._fill_teacher_sheet(ws, lessons, teacher_name, schedule)
        
        wb.save(output_path)
    
    def _fill_teacher_sheet(self, ws, lessons: List, teacher_name: str, schedule):
        """Заполнение листа для преподавателя"""
        # Аналогично группе, но с указанием групп
        ws.merge_cells('A1:F1')
        ws['A1'].value = f'Расписание преподавателя: {teacher_name}'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Шапка
        headers = ['Время'] + self.days[:5]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Время
        for i, time in enumerate(self.times, start=4):
            cell = ws.cell(row=i, column=1)
            cell.value = time
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Матрица
        schedule_matrix = defaultdict(list)
        for lesson in lessons:
            key = (lesson.day, lesson.time_slot)
            schedule_matrix[key].append(lesson)
        
        # Заполнение
        for day in range(5):
            for time_slot in range(7):
                row = 4 + time_slot
                col = 2 + day
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                
                if (day, time_slot) in schedule_matrix:
                    lessons_list = schedule_matrix[(day, time_slot)]
                    cell_value = ""
                    for lesson in lessons_list:
                        cell_value += f"{lesson.subject.name}\n"
                        cell_value += f"Группа: {lesson.group.name}\n"
                        cell_value += f"Ауд. {lesson.room.name}\n"
                    
                    cell.value = cell_value.strip()
                    cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        
        # Размеры
        ws.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 28
        for row in range(4, 11):
            ws.row_dimensions[row].height = 65
    
    def _export_by_rooms(self, schedule, output_path: str):
        """Экспорт по аудиториям"""
        wb = Workbook()
        
        rooms_lessons = defaultdict(list)
        for lesson in schedule.lessons:
            rooms_lessons[lesson.room.name].append(lesson)
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        for room_name, lessons in sorted(rooms_lessons.items()):
            ws = wb.create_sheet(title=f"Ауд. {room_name}"[:31])
            self._fill_room_sheet(ws, lessons, room_name, schedule)
        
        wb.save(output_path)
    
    def _fill_room_sheet(self, ws, lessons: List, room_name: str, schedule):
        """Заполнение листа для аудитории"""
        ws.merge_cells('A1:F1')
        ws['A1'].value = f'Расписание аудитории: {room_name}'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Аналогично преподавателю
        headers = ['Время'] + self.days[:5]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        for i, time in enumerate(self.times, start=4):
            cell = ws.cell(row=i, column=1)
            cell.value = time
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        schedule_matrix = {}
        for lesson in lessons:
            key = (lesson.day, lesson.time_slot)
            schedule_matrix[key] = lesson
        
        for day in range(5):
            for time_slot in range(7):
                row = 4 + time_slot
                col = 2 + day
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                
                if (day, time_slot) in schedule_matrix:
                    lesson = schedule_matrix[(day, time_slot)]
                    cell_value = f"{lesson.subject.name}\n"
                    cell_value += f"Группа: {lesson.group.name}\n"
                    cell_value += f"Преп.: {lesson.teacher.name}"
                    cell.value = cell_value
                    cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        
        ws.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 28
        for row in range(4, 11):
            ws.row_dimensions[row].height = 65
    
    def _export_consolidated(self, schedule, output_path: str):
        """Сводное расписание (все занятия в одной таблице)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводное расписание"
        
        # Заголовок
        ws.merge_cells('A1:F1')
        ws['A1'] = f'СВОДНОЕ РАСПИСАНИЕ - {schedule.name}'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Шапка таблицы
        headers = ['Группа', 'День недели', 'Время', 'Предмет', 'Преподаватель', 'Аудитория']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Сортируем занятия
        sorted_lessons = sorted(
            schedule.lessons,
            key=lambda x: (x.group.name, x.day, x.time_slot)
        )
        
        # Заполняем данные
        for row_idx, lesson in enumerate(sorted_lessons, start=4):
            cells_data = [
                lesson.group.name,
                self.days[lesson.day],
                self.times[lesson.time_slot],
                lesson.subject.name,
                lesson.teacher.name,
                lesson.room.name
            ]
            
            for col_idx, value in enumerate(cells_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(vertical='center')
        
        # Настройка ширины
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 12
        
        wb.save(output_path)